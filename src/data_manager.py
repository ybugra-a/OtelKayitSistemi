"""
Veri Yonetimi - v0.4
Yeni: Rezervasyon, Blacklist, Sirket Ismi, Silme
"""

import os
import json
from datetime import datetime, date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
CONFIG_DIR = os.path.join(BASE_DIR, "config")
EXCEL_PATH = os.path.join(DATA_DIR, "kayitlar.xlsx")
REZERVASYON_PATH = os.path.join(DATA_DIR, "rezervasyonlar.xlsx")
BLACKLIST_PATH = os.path.join(DATA_DIR, "blacklist.xlsx")
CONFIG_PATH = os.path.join(CONFIG_DIR, "ayarlar.cfg")

# Sutun indeksleri (kayitlar.xlsx)
COL_ID = 1
COL_TC = 2
COL_ISIM = 3
COL_SOYISIM = 4
COL_SIRKET = 5
COL_ODA = 6
COL_GIRIS = 7
COL_CIKIS = 8
COL_ODEME_YON = 9
COL_ODEME_TUT = 10
COL_DURUM = 11

DONEMLER = {1:"Q1",2:"Q1",3:"Q1",4:"Q2",5:"Q2",6:"Q2",
            7:"Q3",8:"Q3",9:"Q3",10:"Q4",11:"Q4",12:"Q4"}


class DataManager:
    def __init__(self):
        self._ensure_dirs()
        self._ensure_excel()
        self._ensure_rezervasyon()
        self._ensure_blacklist()
        self._ensure_config()

    def _ensure_dirs(self):
        os.makedirs(DATA_DIR, exist_ok=True)
        os.makedirs(CONFIG_DIR, exist_ok=True)

    # ----------------------------------------------------------------
    # EXCEL OLUSTURMA
    # ----------------------------------------------------------------
    def _ensure_excel(self):
        if not os.path.exists(EXCEL_PATH):
            self._create_new_excel()

    def _create_new_excel(self):
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        now = datetime.now()
        ws = wb.create_sheet(f"{now.year}_{DONEMLER[now.month]}")
        self._create_kayit_headers(ws)
        ws_oda = wb.create_sheet("Odalar")
        ws_oda["A1"] = "Oda Numarasi"
        ws_oda["B1"] = "Durum"
        self._style_header(ws_oda, 1)
        ws_oda.column_dimensions['A'].width = 15
        ws_oda.column_dimensions['B'].width = 15
        for i in range(101, 111):
            ws_oda.append([str(i), "Musait"])
        wb.save(EXCEL_PATH)

    def _create_kayit_headers(self, ws):
        headers = ["Kayit ID","T.C Kimlik No","Isim","Soyisim","Sirket Ismi",
                   "Oda Numarasi","Giris Tarihi","Cikis Tarihi",
                   "Odeme Yontemi","Odeme Tutari","Durum"]
        ws.append(headers)
        self._style_header(ws, 1)
        widths = [12,16,14,14,18,14,14,14,18,14,14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+i)].width = w

    def _ensure_rezervasyon(self):
        if not os.path.exists(REZERVASYON_PATH):
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            ws = wb.create_sheet("Rezervasyonlar")
            headers = ["Rezervasyon ID","T.C Kimlik No","Isim","Soyisim","Sirket Ismi",
                       "Oda Numarasi","Giris Tarihi","Cikis Tarihi",
                       "Odeme Yontemi","Odeme Tutari","Durum","Kayit Tarihi"]
            ws.append(headers)
            self._style_header(ws, 1)
            widths = [14,16,14,14,18,14,14,14,18,14,14,14]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[chr(64+i)].width = w
            wb.save(REZERVASYON_PATH)

    def _ensure_blacklist(self):
        if not os.path.exists(BLACKLIST_PATH):
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            ws = wb.create_sheet("Blacklist")
            headers = ["ID","T.C Kimlik No","Isim","Soyisim","Sirket Ismi","Sebep","Eklenme Tarihi"]
            ws.append(headers)
            self._style_header(ws, 1)
            widths = [10,16,14,14,18,30,16]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[chr(64+i)].width = w
            wb.save(BLACKLIST_PATH)

    def _ensure_config(self):
        if not os.path.exists(CONFIG_PATH):
            self._save_config({"son_yedekleme": None, "version": "1.0"})

    def _style_header(self, ws, row):
        fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
        font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[row]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ----------------------------------------------------------------
    # CONFIG
    # ----------------------------------------------------------------
    def _load_config(self):
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {"son_yedekleme": None, "version": "1.0"}

    def _save_config(self, config):
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

    def get_excel_path(self):
        return EXCEL_PATH

    def check_backup_needed(self):
        config = self._load_config()
        son = config.get("son_yedekleme")
        if not son:
            return True
        try:
            return (date.today() - datetime.strptime(son, "%Y-%m-%d").date()).days > 30
        except:
            return True

    def update_backup_date(self):
        config = self._load_config()
        config["son_yedekleme"] = date.today().strftime("%Y-%m-%d")
        self._save_config(config)

    def get_backup_date(self):
        return self._load_config().get("son_yedekleme", "Hic yedeklenmedi")

    # ----------------------------------------------------------------
    # SHEET / ID YARDIMCILARI
    # ----------------------------------------------------------------
    def get_or_create_sheet(self, year, month):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        name = f"{year}_{DONEMLER[month]}"
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            self._create_kayit_headers(ws)
            wb.save(EXCEL_PATH)
        return name

    def get_next_id(self):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        max_id = 0
        for sn in wb.sheetnames:
            if sn == "Odalar":
                continue
            for row in wb[sn].iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).isdigit():
                    max_id = max(max_id, int(row[0]))
        return max_id + 1

    def _get_next_rezervasyon_id(self):
        wb = openpyxl.load_workbook(REZERVASYON_PATH)
        ws = wb["Rezervasyonlar"]
        max_id = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).isdigit():
                max_id = max(max_id, int(row[0]))
        return max_id + 1

    def _get_next_blacklist_id(self):
        wb = openpyxl.load_workbook(BLACKLIST_PATH)
        ws = wb["Blacklist"]
        max_id = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).isdigit():
                max_id = max(max_id, int(row[0]))
        return max_id + 1

    def _update_oda_durumu(self, wb, oda_no, durum):
        if "Odalar" not in wb.sheetnames:
            return
        for row in wb["Odalar"].iter_rows(min_row=2):
            if str(row[0].value) == str(oda_no):
                row[1].value = durum
                return

    # ----------------------------------------------------------------
    # BLACKLIST
    # ----------------------------------------------------------------
    def blacklist_kontrol(self, tc):
        """TC blacklistte mi? True/False"""
        wb = openpyxl.load_workbook(BLACKLIST_PATH)
        ws = wb["Blacklist"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[1]) == str(tc):
                return True
        return False

    def blacklist_ekle(self, tc, isim, soyisim, sirket, sebep):
        wb = openpyxl.load_workbook(BLACKLIST_PATH)
        ws = wb["Blacklist"]
        bid = self._get_next_blacklist_id()
        ws.append([bid, tc, isim, soyisim, sirket, sebep,
                   date.today().strftime("%d/%m/%Y")])
        wb.save(BLACKLIST_PATH)
        return bid

    def get_blacklist(self, arama=""):
        wb = openpyxl.load_workbook(BLACKLIST_PATH)
        ws = wb["Blacklist"]
        sonuc = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if arama:
                al = arama.lower()
                if (al not in str(row[2]).lower() and
                    al not in str(row[3]).lower() and
                    al not in str(row[1]).lower()):
                    continue
            sonuc.append({
                "id": row[0], "tc": row[1], "isim": row[2],
                "soyisim": row[3], "sirket": row[4],
                "sebep": row[5], "tarih": row[6]
            })
        return sonuc

    def blacklist_sil(self, bl_id):
        wb = openpyxl.load_workbook(BLACKLIST_PATH)
        ws = wb["Blacklist"]
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(bl_id):
                ws.delete_rows(i)
                wb.save(BLACKLIST_PATH)
                return True
        return False

    # ----------------------------------------------------------------
    # REZERVASYON
    # ----------------------------------------------------------------
    def rezervasyon_ekle(self, tc, isim, soyisim, sirket, oda,
                         giris, cikis, odeme_yon, odeme_tut):
        wb = openpyxl.load_workbook(REZERVASYON_PATH)
        ws = wb["Rezervasyonlar"]
        rid = self._get_next_rezervasyon_id()
        giris_str = giris.strftime("%d/%m/%Y") if hasattr(giris, 'strftime') else str(giris)
        cikis_str = cikis.strftime("%d/%m/%Y") if (cikis and hasattr(cikis, 'strftime')) else ""
        odeme_str = str(odeme_tut).replace(',','.') if odeme_tut else ""
        ws.append([rid, tc, isim, soyisim, sirket or "", oda,
                   giris_str, cikis_str, odeme_yon, odeme_str,
                   "Beklemede", date.today().strftime("%d/%m/%Y")])
        wb.save(REZERVASYON_PATH)
        return rid

    def get_rezervasyonlar(self):
        wb = openpyxl.load_workbook(REZERVASYON_PATH)
        ws = wb["Rezervasyonlar"]
        sonuc = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if row[10] == "Iptal":
                continue
            sonuc.append({
                "id": row[0], "tc": row[1], "isim": row[2],
                "soyisim": row[3], "sirket": row[4], "oda": row[5],
                "giris": row[6], "cikis": row[7],
                "odeme_yon": row[8], "odeme_tut": row[9],
                "durum": row[10], "kayit_tarihi": row[11]
            })
        return sonuc

    def get_bekleyen_rezervasyonlar(self):
        """Giris tarihi bugun veya gecmis, hala Beklemede olanlar"""
        bugun = date.today()
        sonuc = []
        for r in self.get_rezervasyonlar():
            if r["durum"] != "Beklemede":
                continue
            try:
                giris_dt = datetime.strptime(str(r["giris"]), "%d/%m/%Y").date()
                if giris_dt <= bugun:
                    sonuc.append(r)
            except:
                pass
        return sonuc

    def rezervasyon_checkin(self, rezervasyon_id):
        """Rezervasyonu aktif kayda donustur"""
        wb_r = openpyxl.load_workbook(REZERVASYON_PATH)
        ws_r = wb_r["Rezervasyonlar"]
        r_data = None
        for row in ws_r.iter_rows(min_row=2):
            if str(row[0].value) == str(rezervasyon_id):
                r_data = [c.value for c in row]
                row[10].value = "Tamamlandi"
                break
        if not r_data:
            return False
        wb_r.save(REZERVASYON_PATH)

        # Ana kayda ekle
        now = datetime.now()
        sheet_name = self.get_or_create_sheet(now.year, now.month)
        kayit_id = self.get_next_id()
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]
        ws.append([kayit_id, r_data[1], r_data[2], r_data[3], r_data[4] or "",
                   r_data[5], r_data[6], r_data[7], r_data[8], r_data[9] or "", "Aktif"])
        self._update_oda_durumu(wb, r_data[5], "Dolu")
        wb.save(EXCEL_PATH)
        return True

    def rezervasyon_iptal(self, rezervasyon_id):
        wb = openpyxl.load_workbook(REZERVASYON_PATH)
        ws = wb["Rezervasyonlar"]
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(rezervasyon_id):
                row[10].value = "Iptal"
                wb.save(REZERVASYON_PATH)
                return True
        return False

    def rezervasyon_sil(self, rezervasyon_id):
        wb = openpyxl.load_workbook(REZERVASYON_PATH)
        ws = wb["Rezervasyonlar"]
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(rezervasyon_id):
                ws.delete_rows(i)
                wb.save(REZERVASYON_PATH)
                return True
        return False

    # ----------------------------------------------------------------
    # KAYIT (Ana)
    # ----------------------------------------------------------------
    def kayit_ekle(self, tc, isim, soyisim, sirket, oda,
                   giris, cikis, odeme_yon, odeme_tut):
        now = datetime.now()
        sheet_name = self.get_or_create_sheet(now.year, now.month)
        kayit_id = self.get_next_id()
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]
        giris_str = giris.strftime("%d/%m/%Y") if hasattr(giris,'strftime') else str(giris)
        cikis_str = cikis.strftime("%d/%m/%Y") if (cikis and hasattr(cikis,'strftime')) else ""
        odeme_str = str(odeme_tut).replace(',','.') if odeme_tut else ""
        ws.append([kayit_id, tc, isim, soyisim, sirket or "",
                   oda, giris_str, cikis_str, odeme_yon, odeme_str, "Aktif"])
        self._update_oda_durumu(wb, oda, "Dolu")
        wb.save(EXCEL_PATH)
        return kayit_id

    def get_aktif_musteriler(self):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sonuc = []
        for sn in wb.sheetnames:
            if sn == "Odalar":
                continue
            for row in wb[sn].iter_rows(min_row=2, values_only=True):
                if row[COL_DURUM-1] == "Aktif":
                    sonuc.append(self._row_to_dict(row, sn))
        return sonuc

    def _row_to_dict(self, row, sheet):
        return {
            "id": row[COL_ID-1], "tc": row[COL_TC-1],
            "isim": row[COL_ISIM-1], "soyisim": row[COL_SOYISIM-1],
            "sirket": row[COL_SIRKET-1],
            "oda": row[COL_ODA-1],
            "giris": row[COL_GIRIS-1], "cikis": row[COL_CIKIS-1],
            "odeme_yon": row[COL_ODEME_YON-1], "odeme_tut": row[COL_ODEME_TUT-1],
            "durum": row[COL_DURUM-1], "sheet": sheet
        }

    def cikis_yaptir(self, kayit_id, sheet_name):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            if str(row[COL_ID-1].value) == str(kayit_id):
                oda = row[COL_ODA-1].value
                row[COL_CIKIS-1].value = datetime.now().strftime("%d/%m/%Y")
                row[COL_DURUM-1].value = "Cikis Yapti"
                self._update_oda_durumu(wb, oda, "Musait")
                wb.save(EXCEL_PATH)
                return True
        return False

    def kayit_guncelle(self, kayit_id, sheet_name, data):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            if str(row[COL_ID-1].value) == str(kayit_id):
                eski_oda = row[COL_ODA-1].value
                yeni_oda = data.get("oda", eski_oda)
                row[COL_TC-1].value = data.get("tc")
                row[COL_ISIM-1].value = data.get("isim")
                row[COL_SOYISIM-1].value = data.get("soyisim")
                row[COL_SIRKET-1].value = data.get("sirket", "")
                row[COL_ODA-1].value = yeni_oda
                row[COL_GIRIS-1].value = data.get("giris")
                row[COL_CIKIS-1].value = data.get("cikis")
                row[COL_ODEME_YON-1].value = data.get("odeme_yon")
                row[COL_ODEME_TUT-1].value = data.get("odeme_tut")
                if str(eski_oda) != str(yeni_oda):
                    self._update_oda_durumu(wb, eski_oda, "Musait")
                    self._update_oda_durumu(wb, yeni_oda, "Dolu")
                wb.save(EXCEL_PATH)
                return True
        return False

    def kayit_sil(self, kayit_id, sheet_name):
        """Kaydi kalici olarak sil, oda musait yap"""
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[COL_ID-1].value) == str(kayit_id):
                oda = row[COL_ODA-1].value
                durum = row[COL_DURUM-1].value
                if durum == "Aktif":
                    self._update_oda_durumu(wb, oda, "Musait")
                ws.delete_rows(i)
                wb.save(EXCEL_PATH)
                return True
        return False

    def get_tum_kayitlar(self, filtre=None):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sonuc = []
        for sn in wb.sheetnames:
            if sn == "Odalar":
                continue
            parts = sn.split("_")
            yil = parts[0] if len(parts) > 0 else ""
            donem = parts[1] if len(parts) > 1 else ""
            for row in wb[sn].iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                k = self._row_to_dict(row, sn)
                k["yil"] = yil
                k["donem"] = donem
                if filtre:
                    arama = filtre.get("arama","").lower()
                    if arama:
                        if (arama not in f"{k['isim']} {k['soyisim']}".lower() and
                            arama not in str(k['tc']).lower() and
                            arama not in str(k.get('sirket','')).lower()):
                            continue
                    if filtre.get("yil","Tumu") != "Tumu" and k["yil"] != filtre["yil"]:
                        continue
                    if filtre.get("donem","Tumu") != "Tumu" and k["donem"] != filtre["donem"]:
                        continue
                    if filtre.get("durum","Tumu") != "Tumu" and k["durum"] != filtre["durum"]:
                        continue
                sonuc.append(k)
        return sonuc

    # ----------------------------------------------------------------
    # ODALAR
    # ----------------------------------------------------------------
    def get_odalar(self):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Odalar" not in wb.sheetnames:
            return []
        sonuc = []
        for row in wb["Odalar"].iter_rows(min_row=2, values_only=True):
            if row[0]:
                sonuc.append({"no": str(row[0]), "durum": str(row[1]) if row[1] else "Musait"})
        return sonuc

    def get_musait_odalar(self):
        return [o["no"] for o in self.get_odalar() if o["durum"] == "Musait"]

    def oda_ekle(self, oda_no):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Odalar" not in wb.sheetnames:
            ws = wb.create_sheet("Odalar")
            ws["A1"] = "Oda Numarasi"
            ws["B1"] = "Durum"
        ws = wb["Odalar"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(oda_no):
                return False
        ws.append([str(oda_no), "Musait"])
        wb.save(EXCEL_PATH)
        return True

    def oda_sil(self, oda_no):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Odalar" not in wb.sheetnames:
            return False
        ws = wb["Odalar"]
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(oda_no):
                ws.delete_rows(i)
                wb.save(EXCEL_PATH)
                return True
        return False

    # ----------------------------------------------------------------
    # AUTOCOMPLETE
    # ----------------------------------------------------------------
    def get_autocomplete_data(self):
        kayitlar = self.get_tum_kayitlar()
        kisiler = {}
        for k in kayitlar:
            key = str(k.get("tc",""))
            if key:
                kisiler[key] = {
                    "tc": k.get("tc",""), "isim": k.get("isim",""),
                    "soyisim": k.get("soyisim",""), "sirket": k.get("sirket",""),
                    "oda": k.get("oda","")
                }
        return list(kisiler.values())

    def get_yillar(self):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        yillar = {sn.split("_")[0] for sn in wb.sheetnames
                  if sn != "Odalar" and "_" in sn}
        return sorted(list(yillar), reverse=True)
